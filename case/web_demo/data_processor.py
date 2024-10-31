import logging
from pathlib import Path
import pandas as pd
import chardet
import logging
import matplotlib.pyplot as plt
from pathlib import Path
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def filter_and_save_order_data(orderNo: str, order_dir: str):
    print(f"[DEBUG] Called filter_and_save_order_data with orderNo: {orderNo}, order_dir: {order_dir}")
    logging.info(f"开始执行数据筛选和保存，订单号: {orderNo}, 保存目录: {order_dir}")
    try:
        # 将 order_dir 转换为 Path 对象
        order_dir_path = Path(order_dir)
        order_dir_path.mkdir(parents=True, exist_ok=True)

        # 尝试多个可能的位置来查找 CSV 文件
        possible_csv_locations = [
            Path(os.getcwd()) / "账务明细.csv",
            Path.home() / "Desktop" / "账务明细.csv",
            Path("C:/Users/Administrator/Desktop/账务明细.csv"),
            # 可以添加更多可能的位置
        ]

        csv_file_path = None
        for location in possible_csv_locations:
            if location.exists():
                csv_file_path = location
                break

        if csv_file_path is None:
            raise FileNotFoundError("无法找到账务明细.csv文件")

        output_csv_path = order_dir_path / f"租金支付流水_{orderNo}.csv"

        logging.info(f"CSV文件路径: {csv_file_path}")
        logging.info(f"输出CSV路径: {output_csv_path}")

        # 这里调用您原来在AdminBusiness类中的filter_and_save_data方法
        result = filter_and_save_data(csv_file_path, orderNo, output_csv_path)

        if result:
            logging.info(f"筛选并保存订单数据步骤完成，订单号: {orderNo}")
        else:
            logging.warning(f"未找到匹配的数据，订单号: {orderNo}")
    except Exception as e:
        logging.error(f"筛选并保存订单数据步骤出错，订单号: {orderNo}，错误: {e}")
        raise


def detect_file_encoding(file_path: Path, num_bytes: int = 10000) -> str:
    """检测文件的编码格式。"""
    try:
        with file_path.open('rb') as f:
            rawdata = f.read(num_bytes)
        result = chardet.detect(rawdata)
        encoding = result['encoding']
        logging.info(f"检测到的编码: {encoding}")
        return encoding
    except Exception as e:
        logging.error(f"检测文件编码时出错: {e}")
        return 'utf-8'  # 默认返回utf-8


def detect_separator(file_path: Path, num_lines: int = 5) -> str:
    """自动检测CSV文件的分隔符。"""
    separators = [',', '\t', ';', '|']
    counts = {sep: 0 for sep in separators}
    try:
        with file_path.open('r', encoding='utf-8', errors='replace') as f:
            for _ in range(num_lines):
                line = f.readline()
                if not line:
                    break
                for sep in separators:
                    counts[sep] += line.count(sep)
        detected_sep = max(counts, key=counts.get)
        logging.info(f"自动检测到的分隔符: '{detected_sep}'")
        return detected_sep
    except Exception as e:
        logging.error(f"自动检测分隔符时出错: {e}")
        return ','  # 默认返回逗号


def read_csv_with_encoding(file_path: Path, encoding: str, sep: str = ',') -> tuple:
    """使用指定编码读取 CSV 文件，保留前4行信息，第5行作为标题，从第6行开始读取数据。"""
    try:
        header_info = pd.read_csv(file_path, encoding=encoding, sep=sep, nrows=4, header=None)
        column_names = \
            pd.read_csv(file_path, encoding=encoding, sep=sep, skiprows=4, nrows=1, header=None).iloc[0]
        df = pd.read_csv(file_path, encoding=encoding, sep=sep, skiprows=5, names=column_names)
        df.columns = df.columns.str.strip().str.replace('\t', '')
        logging.info(f"成功读取 CSV 文件，行数: {df.shape[0]}, 列数: {df.shape[1]}")
        return header_info, df
    except Exception as e:
        logging.error(f"读取 CSV 文件时出错: {e}")
        return None, None


def save_to_csv(df: pd.DataFrame, output_path: Path):
    """保存DataFrame到CSV文件。"""
    try:
        print(f"[DEBUG] Saving DataFrame to {output_path}")
        df.to_csv(output_path, index=False, encoding='utf-8-sig')
        logging.info(f"成功保存数据到 CSV 文件: {output_path}")
        print(f"[DEBUG] Successfully saved CSV to {output_path}")

        # 同时保存为Excel文件
        excel_path = output_path.with_suffix('.xlsx')
        save_to_excel(df, excel_path)
    except PermissionError:
        logging.error(f"保存CSV文件时遇到权限错误: {output_path}")
        alternative_path = Path(os.getcwd()) / output_path.name
        try:
            df.to_csv(alternative_path, index=False, encoding='utf-8-sig')
            logging.info(f"成功保存数据到当前目录的 CSV 文件: {alternative_path}")
            print(f"[DEBUG] Successfully saved CSV to alternative path: {alternative_path}")

            # 同时保存为Excel文件
            excel_path = alternative_path.with_suffix('.xlsx')
            save_to_excel(df, excel_path)
        except Exception as e:
            logging.error(f"保存到当前目录的CSV文件时也出错: {e}")
            print(f"[DEBUG] Failed to save CSV to alternative path: {alternative_path}, error: {e}")
    except Exception as e:
        logging.error(f"保存数据到 CSV 时出错: {e}")
        print(f"[DEBUG] Error saving CSV to {output_path}: {e}")


def save_to_excel(df: pd.DataFrame, output_path: Path):
    """保存DataFrame到Excel文件并调整列宽，删除中间生成的CSV文件。"""
    try:
        # 如果Excel文件已存在，先删除
        if output_path.exists():
            os.remove(output_path)
            logging.info(f"删除了旧的 Excel 文件: {output_path}")

        wb = Workbook()
        ws = wb.active

        # 写入列名
        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column_title

        # 写入数据
        for row_num, row in enumerate(df.values, 2):
            for col_num, cell_value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = cell_value

        # 调整列宽
        for col_num, column in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_num)
            column_title = ws.cell(row=1, column=col_num).value

            print(f"[DEBUG] 处理列: {column_title}")

            for cell in column:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass

            print(f"[DEBUG] 列 {column_title} 的最大长度: {max_length}")

            # 基础宽度调整
            adjusted_width = max_length + 4

            # 对特定列进行额外的宽度调整
            if '金额' in column_title or '日期' in column_title:
                adjusted_width = max(adjusted_width, 20)
            elif '对方账号' in column_title:
                adjusted_width = max(adjusted_width, 40)  # 为"对方账号"列设置更大的宽度
            elif '商品名称' in column_title or '备注' in column_title:
                adjusted_width = max(adjusted_width, 30)  # 为可能包含长文本的列设置更大的宽度
            else:
                adjusted_width = max(adjusted_width, 15)  # 其他列的最小宽度

            print(f"[DEBUG] 列 {column_title} 的最终宽度: {adjusted_width}")

            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(output_path)
        logging.info(f"成功保存数据到新的 Excel 文件并调整列宽: {output_path}")
        print(f"[DEBUG] Successfully saved new Excel to {output_path}")

    except Exception as e:
        logging.error(f"保存数据到 Excel 时出错: {e}")
        print(f"[DEBUG] Error saving Excel to {output_path}: {e}")
        raise  # 重新抛出异常，因为保存 Excel 是关键步骤
    
# 单独处理删除 CSV 文件
    try:
        # 删除中间生成的CSV文件
        csv_path = output_path.with_suffix('.csv')
        if csv_path.exists():
            os.remove(csv_path)
            logging.info(f"删除了中间生成的 CSV 文件: {csv_path}")
        else:
            logging.warning(f"未找到中间生成的 CSV 文件: {csv_path}")
    except Exception as e:
        logging.error(f"删除中间生成的 CSV 文件时出错: {e}")
        print(f"[DEBUG] Error deleting CSV file {csv_path}: {e}")



def filter_and_save_data(file_path: Path, order_no: str, output_csv: Path):
    """筛选包含指定订单号的数据并保存。"""
    try:
        order_no_str = str(order_no)
        logging.info(f"筛选订单号（字符串）: {order_no_str}")

        encoding = detect_file_encoding(file_path)
        sep = detect_separator(file_path)

        header_info, df = read_csv_with_encoding(file_path, encoding, sep)
        if df is None:
            for alt_enc in ['gbk', 'gb18030', 'utf-8']:
                logging.info(f"尝试使用替代编码 {alt_enc} 读取 CSV 文件")
                header_info, df = read_csv_with_encoding(file_path, alt_enc, sep)
                if df is not None:
                    break
            if df is None:
                logging.error("所有编码尝试均失败，无法读取 CSV 文件")
                raise ValueError("无法读取 CSV 文件，所有编码尝试均失败。")

        # 找出所有可能包含订单号的列
        potential_columns = [col for col in df.columns if '订单号' in col or col == '备注']
        logging.info(f"可能包含订单号的列: {potential_columns}")

        if potential_columns:
            logging.info(f"使用以下列进行过滤: {', '.join(potential_columns)}")

            # 创建过滤条件
            filter_condition = pd.Series(False, index=df.index)
            for col in potential_columns:
                filter_condition |= df[col].astype(str).str.contains(order_no_str, na=False)

            filtered_df = df[filter_condition]
        else:
            logging.warning("未找到可能包含订单号的列，使用包含方式在所有列中进行过滤")
            filtered_df = df[df.apply(lambda row: order_no_str in ' '.join(row.astype(str)), axis=1)]

        logging.info(f"筛选完成，找到 {len(filtered_df)} 条记录")
        print(f"[DEBUG] 筛选完成，找到 {len(filtered_df)} 条记录")

        if len(filtered_df) > 0:
            logging.info("筛选出的数据：")
            logging.info(filtered_df.to_string())
            save_to_csv(filtered_df, output_csv)
            logging.info(f"筛选后的数据已保存至 {output_csv}")
            print(f"[DEBUG] Data saved to {output_csv}")
            return True
        else:
            logging.warning(f"没有找到包含订单号 {order_no_str} 的行，跳过保存步骤")
            print(f"[DEBUG] 没有找到包含订单号 {order_no_str} 的行")
            return False

    except FileNotFoundError:
        logging.error(f"找不到文件: {file_path}")
        raise
    except PermissionError:
        logging.error(f"没有权限访问文件: {file_path}")
        raise
    except pd.errors.EmptyDataError:
        logging.error(f"CSV 文件为空: {file_path}")
        raise
    except Exception as e:
        logging.error(f"处理文件时发生未知错误: {e}")
        raise